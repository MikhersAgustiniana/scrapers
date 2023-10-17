import requests
import openpyxl

def limpieza(response):
	response = response.json()
	del response["effective_url"]
	del response["word_count"]
	del response["og_url"]
	del response["og_image"]
	del response["og_type"]
	del response["twitter_site"]
	del response["twitter_creator"]
	del response["twitter_image"]
	del response["twitter_title"]
	del response["twitter_description"]
	del response["twitter_card"]
	return response

def txtfly(url):
	api = "https://full-text-rss.p.rapidapi.com/extract.php"
	payload = {
		"url": f"{url}",
		"xss": "1",
		"lang": "2",
		"links": "remove",
		"content": "text"
	}
	headers = {
		"content-type": "application/x-www-form-urlencoded",
		"X-RapidAPI-Key": "",
		"X-RapidAPI-Host": "full-text-rss.p.rapidapi.com"
	}
	data = requests.post(api, data=payload, headers=headers)
	data_limpia = limpieza(data)
	return data_limpia

workbook = openpyxl.Workbook()
sheet = workbook.active

urls = ["https://www.eltiempo.com/bogota/carros-abandonados-generan-un-deficit-de-184-000-millones-por-que-815700",
        "https://www.eltiempo.com/justicia/investigacion/por-inaccion-de-fiscalia-jineth-bedoya-desiste-de-investigacion-por-amenazas-que-sufrio-815878"]
lista_scraper = []
for url in urls:
      lista_scraper.append(txtfly(url))
      
columnas = list(lista_scraper[0].keys())
for col_idx, columna in enumerate(columnas, start=1):
    sheet.cell(row=1, column=col_idx, value=columna)

for row_idx, objeto in enumerate(lista_scraper, start=2):
    for col_idx, columna in enumerate(columnas, start=1):
        sheet.cell(row=row_idx, column=col_idx, value=objeto[columna])

workbook.save("scrapers.xlsx")